from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from openpyxl import Workbook, load_workbook

from ..database import get_db
from ..dependencies import get_current_user, require_role
from ..schemas.review_type import ReviewTypeCreate, ReviewTypeUpdate, ReviewTypeStatusUpdate
from ..models.review_type import ReviewType
from ..models.template import Template
from ..models.template_version import TemplateVersion
from ..models.user import User

router = APIRouter()


@router.get("")
async def list_review_types(
    include_inactive: bool = Query(True),
    name: Optional[str] = Query(None, description="按名称模糊查询"),
    project_category: Optional[str] = Query(None, description="按项目细类筛选"),
    sub_category: Optional[str] = Query(None, description="按二级分类筛选"),
    status: Optional[str] = Query(None, description="按状态筛选"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    query = select(ReviewType).order_by(ReviewType.created_at.desc())
    if not include_inactive:
        query = query.where(ReviewType.status == "active")
    if name:
        query = query.where(ReviewType.name.contains(name))
    if project_category:
        query = query.where(ReviewType.project_category == project_category)
    if sub_category:
        query = query.where(ReviewType.sub_category == sub_category)
    if status:
        query = query.where(ReviewType.status == status)
    
    result = await db.execute(query)
    types = result.scalars().all()
    
    items = []
    for rt in types:
        item = rt.to_dict()
        # 有效模板数：非失效 + 当前版本状态为 published 的模板
        active_count_result = await db.execute(
            select(func.count()).select_from(Template).join(
                TemplateVersion, Template.current_version_id == TemplateVersion.id
            ).where(
                Template.review_type_id == rt.id,
                Template.status != "archived",
                TemplateVersion.status == "published"
            )
        )
        item["active_template_count"] = active_count_result.scalar()

        # 已失效模板数：已失效的模板
        archived_result = await db.execute(
            select(func.count()).select_from(Template).where(
                Template.review_type_id == rt.id,
                Template.status == "archived"
            )
        )
        item["expired_template_count"] = archived_result.scalar()

        items.append(item)
    
    return {"items": items}


@router.get("/{type_id}")
async def get_review_type(
    type_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(ReviewType).where(ReviewType.id == type_id))
    rt = result.scalar_one_or_none()
    if not rt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="评审阶段不存在")
    return rt.to_dict()


@router.post("")
async def create_review_type(
    data: ReviewTypeCreate,
    current_user: User = Depends(require_role("sys_admin")),
    db: AsyncSession = Depends(get_db),
):
    rt = ReviewType(
        name=data.name,
        project_category=data.project_category,
        sub_category=data.sub_category,
        created_by=current_user.id,
    )
    db.add(rt)
    await db.flush()
    await db.refresh(rt)
    return rt.to_dict()


@router.put("/{type_id}")
async def update_review_type(
    type_id: int,
    data: ReviewTypeUpdate,
    current_user: User = Depends(require_role("sys_admin")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(ReviewType).where(ReviewType.id == type_id))
    rt = result.scalar_one_or_none()
    if not rt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="评审阶段不存在")
    
    if data.name is not None:
        rt.name = data.name
    if data.project_category is not None:
        rt.project_category = data.project_category
    if data.sub_category is not None:
        rt.sub_category = data.sub_category

    db.add(rt)
    await db.flush()
    await db.refresh(rt)
    return rt.to_dict()


@router.delete("/{type_id}")
async def delete_review_type(
    type_id: int,
    current_user: User = Depends(require_role("sys_admin")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(ReviewType).where(ReviewType.id == type_id))
    rt = result.scalar_one_or_none()
    if not rt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="评审阶段不存在")
    
    # Check if templates exist
    count_result = await db.execute(
        select(func.count()).select_from(Template).where(Template.review_type_id == type_id)
    )
    if count_result.scalar() > 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="该评审阶段下存在模板，无法删除")
    
    await db.delete(rt)
    await db.flush()
    return {"message": "评审阶段已删除"}


@router.put("/{type_id}/status")
async def update_review_type_status(
    type_id: int,
    data: ReviewTypeStatusUpdate,
    current_user: User = Depends(require_role("sys_admin")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(ReviewType).where(ReviewType.id == type_id))
    rt = result.scalar_one_or_none()
    if not rt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="评审阶段不存在")
    
    rt.status = data.status
    db.add(rt)
    await db.flush()
    await db.refresh(rt)
    return rt.to_dict()


@router.get("/export/template")
async def download_import_template():
    """下载评审阶段导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "评审阶段导入模板"
    ws.append(["评审名称", "项目细类", "二级分类", "状态"])
    # 示例数据
    ws.append(["发动机评审", "产品类", "发动机", "启用"])
    # 列宽
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=review_type_template.xlsx"},
    )


@router.post("/import")
async def import_review_types(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role("sys_admin")),
    db: AsyncSession = Depends(get_db),
):
    """Excel 批量导入评审阶段"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 文件")

    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    # 映射
    category_map = {"产品类": "product", "工艺类": "process", "研究类": "research"}
    status_map = {"启用": "active", "停用": "inactive"}

    created = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not row or not row[0] or str(row[0]).strip() == "":
            continue

        name = str(row[0]).strip()
        category_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        sub_category = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        status_raw = str(row[3]).strip() if len(row) > 3 and row[3] else "启用"

        category = category_map.get(category_raw, category_raw if category_raw else None)
        status = status_map.get(status_raw, "active")

        # 检查是否已存在同名
        existing = await db.execute(
            select(ReviewType).where(ReviewType.name == name)
        )
        if existing.scalar_one_or_none():
            errors.append(f"第{i}行: 「{name}」已存在，跳过")
            continue

        try:
            rt = ReviewType(
                name=name,
                project_category=category,
                sub_category=sub_category or None,
                status=status,
                created_by=current_user.id,
            )
            db.add(rt)
            created += 1
        except Exception as e:
            errors.append(f"第{i}行: {str(e)}")

    await db.flush()

    return {
        "created": created,
        "errors": errors,
        "message": f"成功导入 {created} 条" + (f"，{len(errors)} 条跳过" if errors else ""),
    }
